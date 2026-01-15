from __future__ import annotations

import re
from pathlib import Path
from typing import Dict, List, Tuple

import pandas as pd
from openpyxl import load_workbook


def _normalize(s: str) -> str:
    return str(s).strip()


def build_replacements_map(abbrev_path: str, sheet_col="Sheet", code_col="Code", desc_col="Description"):
    """
   Build a per-sheet abbreviation replacement map from an Excel file.

   The abbreviations Excel file is expected to contain at least these columns:
   - sheet_col (default: "Sheet"): worksheet/category name in the target workbook
   - code_col  (default: "Code"): abbreviation/code to be replaced
   - desc_col  (default: "Description"): expanded text to replace the code with

   Behavior:
   - Trims whitespace from Code/Description/Sheet fields.
   - Drops rows where Code or Sheet is empty.
   - Groups rows by Sheet and produces a mapping:
        replacements_by_sheet[sheet_name] = [(code, description), ...]
   - If the same Code appears multiple times within the same Sheet with different Descriptions,
      a warning is recorded and the first occurrence is used.
   - Sorts replacement pairs by code length (longer first) to reduce partial-overlap issues
      (e.g., replacing "ATR" before "AT").

   Returns:
        replacements_by_sheet: dict[str, list[tuple[str, str]]]
            Mapping of worksheet name -> list of (code, description) replacements.
        warnings: list[str]
            Human-readable warnings about ambiguous codes and other issues.
    """
    df = pd.read_excel(abbrev_path)

    # Basic cleanup
    df = df[[code_col, desc_col, sheet_col]].copy()
    df[code_col] = df[code_col].astype(str).map(_normalize)
    df[desc_col] = df[desc_col].astype(str).map(_normalize)
    df[sheet_col] = df[sheet_col].astype(str).map(_normalize)

    # Drop empty codes/sheets
    df = df[(df[code_col] != "") & (df[sheet_col] != "")]

    warnings: List[str] = []
    replacements_by_sheet: Dict[str, List[Tuple[str, str]]] = {}

    # Group by sheet
    for sheet_name, g in df.groupby(sheet_col, dropna=False):
        # Detect duplicates within the same sheet
        dup = g.groupby(code_col)[desc_col].nunique()
        ambiguous_codes = dup[dup > 1].index.tolist()
        if ambiguous_codes:
            warnings.append(
                f"Ambiguous codes in sheet '{sheet_name}': {ambiguous_codes} "
                f"(same Code mapped to multiple Descriptions). Using the first occurrence."
            )

        # Keep first mapping for duplicates
        g2 = g.drop_duplicates(subset=[code_col], keep="first")

        pairs = list(zip(g2[code_col].tolist(), g2[desc_col].tolist()))

        # Important: replace longer codes first to avoid partial overlaps (e.g., AT vs ATR)
        pairs.sort(key=lambda x: len(x[0]), reverse=True)

        replacements_by_sheet[sheet_name] = pairs

    return replacements_by_sheet, warnings


def replace_in_workbook(
    items_path: str,
    abbrev_path: str,
    output_path: str,
    match_whole_tokens: bool = True,
    case_sensitive: bool = True,
):
    """
    Scans ALL cells in the target sheet(s) and replaces Code -> Description.
    Only processes worksheets that appear in the abbreviations file.
    """
    replacements_by_sheet, warnings = build_replacements_map(abbrev_path)

    wb = load_workbook(items_path)
    for sheet_name, pairs in replacements_by_sheet.items():
        if sheet_name not in wb.sheetnames:
            warnings.append(f"Sheet '{sheet_name}' not found in items workbook. Skipping.")
            continue

        ws = wb[sheet_name]

        # Build regex patterns for this sheet
        compiled: List[Tuple[re.Pattern, str]] = []
        flags = 0 if case_sensitive else re.IGNORECASE

        for code, desc in pairs:
            # If you want "AH" to match as a token, not inside other words:
            # - match_whole_tokens=True uses boundary-like logic:
            #   not preceded/followed by letter/number/underscore.
            if match_whole_tokens:
                pattern = rf"(?<![A-Za-z0-9_]){re.escape(code)}(?![A-Za-z0-9_])"
            else:
                pattern = re.escape(code)

            compiled.append((re.compile(pattern, flags), desc))

        # Fast exact-lookup map (code -> description)
        exact_map = {code: desc for code, desc in pairs}

        # Iterate through ALL cells
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if v is None:
                    continue

                # Case 1: string cells (existing behavior)
                if isinstance(v, str) and v.strip():
                    original = v
                    new_val = original
                    for pat, desc in compiled:
                        new_val = pat.sub(desc, new_val)
                    if new_val != original:
                        cell.value = new_val
                    continue

                # Case 2: numeric cells (NEW)
                # Only replace if the ENTIRE cell matches a code 100%
                if isinstance(v, (int, float)):
                    # normalize 7948.0 -> "7948"
                    if isinstance(v, float) and v.is_integer():
                        key = str(int(v))
                    else:
                        key = str(v)

                    if key in exact_map:
                        cell.value = exact_map[key]

    wb.save(output_path)
    return warnings


if __name__ == "__main__":
    # Example usage
    items_file = "data/excel_files/item_logic_abbreviation.xlsx"
    abbrev_file = "/Users/sufarhansudaryo/Documents/Work/semantic-automapper/data/excel_files/Abbreviation_expanded.xlsx"
    out_file = "items_replaced.xlsx"

    warnings = replace_in_workbook(
        items_path=items_file,
        abbrev_path=abbrev_file,
        output_path=out_file,
        match_whole_tokens=True,   # recommended
        case_sensitive=True,       # set False if items might use "ah" vs "AH"
    )

    if warnings:
        print("\nWARNINGS:")
        for w in warnings:
            print("-", w)

    print(f"\nDone. Saved: {out_file}")
