import openpyxl
import argparse
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

"""
Excel Color-Coding for Top-1/Top-2/Top-3 Classification Matches
---------------------------------------------------------------

This script loads an Excel sheet, compares a mapping column to the Top-1/Top-2/Top-3
prediction columns, colors each row based on the highest correct match, and generates
a summary sheet with basic accuracy statistics.

Process:
1. Validate required columns (Top-1, Top-2, Top-3, Mapping).
2. Color rows:
   - Top-1 match → green
   - Top-2 match → yellow
   - Top-3 match → orange
   - Mismatch → red
   - No mapping → no fill
3. Create a “Summary” sheet with match counts and percentages.

Output:
A new Excel file containing:
 - Color-coded rows
 - A summary sheet with Top-1/2/3 accuracy

Example usage:
--------------
python validator.py \
  --input_file predictions.xlsx \
  --sheet_name Sheet1 \
  --top1_column top_1_class_name \
  --top2_column top_2_class_name \
  --top3_column top_3_class_name \
  --mapping_column mapping_ph \
  --output_file predictions_colored.xlsx
"""


# Color fills
FILLS = {
    "top1": PatternFill("solid", fgColor="C6EFCE"),    # green
    "top2": PatternFill("solid", fgColor="FFEB9C"),    # yellow
    "top3": PatternFill("solid", fgColor="F8CBAD"),    # orange
    "mismatch": PatternFill("solid", fgColor="FFC7CE"),  # red
    "no_mapping": PatternFill(fill_type=None)           # no fill
}



def load_sheet(input_file, sheet_name):
    wb = openpyxl.load_workbook(input_file)
    ws = wb[sheet_name]
    return wb, ws


def get_column_indices(ws, top1, top2, top3, mapping):
    header = {cell.value: cell.column for cell in ws[1]}

    required = [top1, top2, top3, mapping]
    for col in required:
        if col not in header:
            raise ValueError(f"Column '{col}' not found.")

    return (
        header[top1],
        header[top2],
        header[top3],
        header[mapping]
    )


def color_rows(ws, col1, col2, col3, col_map):
    stats = {"total": 0, "mapping": 0, "top1": 0, "top2": 0, "top3": 0}

    for row in ws.iter_rows(min_row=2):
        stats["total"] += 1

        mapping = row[col_map - 1].value
        top1 = row[col1 - 1].value
        top2 = row[col2 - 1].value
        top3 = row[col3 - 1].value

        if not mapping:  
            fill = FILLS["no_mapping"]
        else:
            stats["mapping"] += 1

            if top1 == mapping:
                stats["top1"] += 1
                fill = FILLS["top1"]

            elif top2 == mapping:
                stats["top2"] += 1
                fill = FILLS["top2"]

            elif top3 == mapping:
                stats["top3"] += 1
                fill = FILLS["top3"]

            else:
                fill = FILLS["mismatch"]

        for c in row:
            c.fill = fill

    return stats


def create_summary_sheet(wb, stats):
    summary = wb.create_sheet("Summary")

    mapping = stats["mapping"]

    summary.append(["Metric", "Value"])
    summary.append(["Total Items", stats["total"]])
    summary.append(["Total with Mapping", stats["mapping"]])
    summary.append([])

    if mapping > 0:
        pct1 = stats["top1"] / mapping
        pct2 = stats["top2"] / mapping
        pct3 = stats["top3"] / mapping
        total_match = (stats["top1"] + stats["top2"] + stats["top3"]) / mapping
        false_match = 1 - total_match

        summary.append(["Top-1 Match %", f"{pct1*100:.2f}%"])
        summary.append(["Top-2 Match %", f"{pct2*100:.2f}%"])
        summary.append(["Top-3 Match %", f"{pct3*100:.2f}%"])
        summary.append(["Total Match % (Top1-3)", f"{total_match*100:.2f}%"])
        summary.append(["False Match / No Match %", f"{false_match*100:.2f}%"])
    else:
        summary.append(["Top-1 Match %", "N/A"])
        summary.append(["Top-2 Match %", "N/A"])
        summary.append(["Top-3 Match %", "N/A"])
        summary.append(["Total Match %", "N/A"])
        summary.append(["False Match %", "N/A"])

    # Auto-size columns
    for col in summary.columns:
        width = max(len(str(c.value)) if c.value else 0 for c in col)
        summary.column_dimensions[get_column_letter(col[0].column)].width = width + 4



def main():
    parser = argparse.ArgumentParser(description="Color-code Excel rows based on Top-1/Top-2/Top-3 match and produce summary.")

    parser.add_argument("--input_file", required=True)
    parser.add_argument("--sheet_name", required=True)
    parser.add_argument("--top1_column", required=True)
    parser.add_argument("--top2_column", required=True)
    parser.add_argument("--top3_column", required=True)
    parser.add_argument("--mapping_column", required=True)
    parser.add_argument("--output_file", required=True)

    args = parser.parse_args()

    wb, ws = load_sheet(args.input_file, args.sheet_name)

    col1, col2, col3, col_map = get_column_indices(
        ws, args.top1_column, args.top2_column, args.top3_column, args.mapping_column
    )

    stats = color_rows(ws, col1, col2, col3, col_map)
    create_summary_sheet(wb, stats)

    wb.save(args.output_file)
    print(f"✔ Output saved as: {args.output_file}")


if __name__ == "__main__":
    main()